# ========== Imports ==========
from customtkinter import *
import tkinter as tk
from PIL import ImageTk, Image
from openpyxl import Workbook, load_workbook
import os

# ========== Main Window Setup ==========
main_window = CTk()
main_window.withdraw() # To hide the main window temporarily

# ========== Window Dimensions Setup ==========
screen_width = main_window.winfo_screenwidth()
screen_height = main_window.winfo_screenheight()

window_width = int(screen_width * 0.7)
window_height = int(screen_height * 0.7)

min_width = 850
min_height = 600

window_width = max(window_width, min_width)
window_height = max(window_height, min_height)

x_position = (screen_width - window_width) // 2
y_position = (screen_height - window_height) // 2

# ========== Excel File Setup ==========
EXCEL_FILE = "accounts.xlsx"

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["username", "password"])
    ws.append(["admin", "1234"])  # Default admin account
    wb.save(EXCEL_FILE)

# ==========================================================
# ===================== FUNCTIONS ==========================
# ==========================================================

# ========== Check if account exists Function ==========
def account_exists(username):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return True
    return False

# ========== Validate login Function ==========
def validate_login(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

# ========== Add account Function ==========
def add_account(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([username, password])
    wb.save(EXCEL_FILE)

# ========== Attempt login Function ==========
def attempt_login():
    username = username_entry.get()
    password = password_entry.get()
    if validate_login(username, password):
        login_window.destroy()
        main_window.deiconify()  # To show the main window after successful login
    else:
        error_label.configure(text="Invalid credentials", text_color="red")

# ========== Attempt sign up Function ==========
def attempt_signup():
    username = username_entry.get()
    password = password_entry.get()
    if not username or not password:
        error_label.configure(text="Please enter both username and password", text_color="red")
        return
    if account_exists(username):
        error_label.configure(text="Username already exists", text_color="red")
        return
    add_account(username, password)
    error_label.configure(text="Account created! Please login.", text_color="green")

# ========== Focus password entry Function ==========
def focus_password(event):
    password_entry.focus_set()

# ========== Submit login on enter key Function ==========
def submit_login(event):
    attempt_login()

# ========== Toggle password visibility Function ==========
def toggle_password():
    if password_entry.cget("show") == "":
        password_entry.configure(show = "*")
        toggle_button.configure(text = "Show")
    else:
        password_entry.configure(show = "")
        toggle_button.configure(text = "Hide")

# ========== Logout Function ==========
def logout():
    main_window.withdraw()  # Hide the main window
    login()

# ========== Set window icon Function ==========
def set_window_icon(window):
    try:
        window.iconbitmap("sts_icon.ico")
    except Exception as e:
        print(f"Failed to set icon: {e}")

# ========== Login window Function ==========
def login():
    global login_window, username_entry, password_entry, error_label, toggle_button
    login_window = tk.Toplevel()
    login_window.title("Sales Tracker | Login")
    login_window.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
    login_window.resizable(0, 0)
    set_window_icon(login_window)
    login_window.config(bg="#242424")

    # ========== Side Image Setup ==========
    side_img = Image.open("side-img.png")
    logo = Image.open("sales-icon-20.png")

    side_img = CTkImage(dark_image=side_img, light_image=side_img, size=(window_width // 2, window_height))
    logo = CTkImage(dark_image=logo, light_image=logo, size=(35, 35))

    image_label = CTkLabel(login_window, image=side_img, text="")
    image_label.pack(side="left")

    # ========== Login Frame Setup ==========
    login_frame = CTkFrame(login_window, fg_color="#272d3c", border_color="#0b121e", border_width=2)
    login_frame.pack(side="left", fill="both", expand=True)

    CTkLabel(login_frame, image=logo, text=" SALES TRACKER SYSTEM", compound="left", text_color="#f55247", font=("Arial", 25, "bold")).pack(pady=(100, 90))

    CTkLabel(login_frame, text="Username", font=("Arial", 15)).pack(anchor="w", padx=100)
    username_entry = CTkEntry(login_frame, width=window_width // 5 - 10, height=30)
    username_entry.pack(anchor="w", padx=130, pady=5)

    CTkLabel(login_frame, text="Password", font=("Arial", 15)).pack(anchor="w", padx=100, pady=(15, 0))

    # ========== Password Entry and Toggle Button ==========
    password_frame = CTkFrame(login_frame, fg_color="transparent")
    password_frame.pack(anchor="w", padx=(130, 0), pady=5)

    password_entry = CTkEntry(password_frame, width=window_width // 5 - 10, height=30, show="*")
    password_entry.pack(side="left")

    toggle_button = CTkButton(password_frame, text="Show", command=toggle_password, width=50)
    toggle_button.pack(side="left", padx=(5, 0)) 

    # ========== Entry Bindings ==========
    username_entry.bind("<Return>", focus_password)
    password_entry.bind("<Return>", submit_login)

    # ========== Sign Up and Login Buttons ==========
    signup_login_frame = CTkFrame(login_frame, fg_color="transparent")
    signup_login_frame.pack(anchor="w", padx=130, pady=20)

    CTkButton(signup_login_frame, text="Sign Up", font=("Arial", 13), command=attempt_signup, fg_color="#f55247", width=75, height=30).pack(side="left", padx=(5, 20))
    CTkButton(signup_login_frame, text="Login", font=("Arial", 13), command=attempt_login, fg_color="#2d8cff", width=75, height=30).pack(side="right")

    # ========== Error Label ==========
    error_label = CTkLabel(login_frame, text="")
    error_label.pack()

# ==========================================================
# ===================== UI SETUP ===========================
# ==========================================================

# ========== Login Window ==========
login()

# ========== Main Window UI ==========
main_window.title("Sales Tracker | Home")
main_window.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
set_window_icon(main_window)

CTkLabel(main_window, text="Welcome to Sales Tracker!", font=("Arial", 20)).pack(pady=40)
CTkButton(main_window, text="Logout", command=logout).pack(pady=20)
CTkButton(main_window, text="Quit", command=main_window.destroy).pack(pady=20)

# ========== Main Loop ==========
main_window.mainloop()